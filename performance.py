import os.path

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook


def smape(y_true, y_pred):
    """
    Compute SMAPE（Symmetric Mean Absolute Percentage Error）
    """
    try:
        n = len(y_true)
        numerator = np.abs(y_pred - y_true)
        denominator = (np.abs(y_true) + np.abs(y_pred))  # / 2
        smape = np.sum(numerator / denominator) * (100.0 / n)
        return smape
    except:
        return -1


def is_sheet_exists(file_path, sheet_name):
    return sheet_name in pd.read_excel(file_path, sheet_name=None)


def save_performance(output_dir, region, rmse, rmse_normalized,
                     mae, mae_normalized,
                     smape, r2, model, index_of_dataset_begin, index_of_dataset_end,
                     index_of_train_begin, index_of_train_end,
                     index_of_test_begin, index_of_test_end, ahead=0):
    # Load existing Excel file or create a new one if it doesn't exist
    output_file = f'{output_dir}/performance.xlsx'
    if os.path.exists(output_file):
        df = pd.read_excel(output_file)
    else:
        df = pd.DataFrame()

    new_row = {
        'Region': region,
        'RMSE': rmse,
        'RMSE_normalized': rmse_normalized,
        'MAE': mae,
        'MAE_normalized': mae_normalized,
        'SMAPE': smape,
        'R2': r2,
        'Model': model,
        'Train_start': index_of_train_begin,
        'Train_end': index_of_train_end,
        'Test_start': index_of_test_begin,
        'Test_end': index_of_test_end,
        'Ahead': ahead + 1
    }

    df = pd.concat([df, pd.DataFrame(new_row, index=[0])])
    print(f'save_performance: df={df}')
    if os.path.exists(output_file):
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False)
    else:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False)


    # try:
    #     wb = load_workbook(file_path)
    #     ws = wb.active
    # except FileNotFoundError:
    #     wb = Workbook()
    #     ws = wb.active
    #     ws.append(['Region', 'RMSE', 'RMSE_normalized',
    #                'MAE', 'MAE_normalized',
    #                'SMAPE', 'R2', 'Model',
    #                'Train_start', 'Train_end', 'Test_start', 'Test_end', 'Ahead'])
    # ws.append([region,
    #            rmse, rmse_normalized,
    #            mae, mae_normalized,
    #            smape, r2, model,
    #            index_of_train_begin, index_of_train_end,
    #            index_of_test_begin, index_of_test_end, ahead + 1])
    # wb.save(file_path)
